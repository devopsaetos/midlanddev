# -*- coding: utf-8 -*-
import base64
import io

from openpyxl import load_workbook

from odoo import models, fields, _
from odoo.exceptions import UserError, ValidationError


class UnitImportWizard(models.TransientModel):
    _name = 'land.development.unit.import'
    _description = 'Import Units (Building)'

    file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='File Name')

    def action_import(self):
        self.ensure_one()

        try:
            workbook = load_workbook(io.BytesIO(base64.b64decode(self.file)), data_only=True)
        except Exception:
            raise UserError(_("Could not read the file. Please upload a valid .xlsx file exported from the Unit list."))

        sheet = workbook.active
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        if not rows:
            raise UserError(_("The file has no data rows."))

        Unit = self.env['plot.inventory']
        state_map = {str(label).strip().lower(): key for key, label in Unit._fields['state'].selection}
        possession_map = {
            str(label).strip().lower(): key for key, label in Unit._fields['possession_status'].selection
        }

        updated = 0
        errors = []
        for index, row in enumerate(rows, start=2):
            def cell(col):
                return row[col] if col < len(row) else None

            serial_number = cell(0)
            if not serial_number:
                continue
            serial_number = str(serial_number).strip()

            unit = Unit.search([
                ('serial_number', '=', serial_number),
                ('project_type', '=', 'skyscraper'),
            ], limit=1)
            if not unit:
                errors.append(_("Row %s: no Building unit found with Serial Number '%s'") % (index, serial_number))
                continue

            vals = {}

            unit_number = cell(1)
            if unit_number is not None and str(unit_number).strip():
                vals['name'] = str(unit_number).strip()

            for col, field_name in ((9, 'net_area'), (10, 'balcony_area'), (11, 'standard_area'), (12, 'actual_area')):
                value = cell(col)
                if value is None or value == '':
                    continue
                try:
                    vals[field_name] = float(value)
                except (TypeError, ValueError):
                    errors.append(_("Row %s: invalid number for column '%s'") % (index, field_name))

            state_value = cell(13)
            if state_value:
                key = state_map.get(str(state_value).strip().lower())
                if key:
                    vals['state'] = key
                else:
                    errors.append(_("Row %s: unknown State '%s'") % (index, state_value))

            possession_value = cell(14)
            if possession_value:
                key = possession_map.get(str(possession_value).strip().lower())
                if key:
                    vals['possession_status'] = key
                else:
                    errors.append(_("Row %s: unknown Possession Status '%s'") % (index, possession_value))

            if vals:
                try:
                    with self.env.cr.savepoint():
                        unit.write(vals)
                    updated += 1
                except ValidationError as e:
                    errors.append(_("Row %s: %s") % (index, e.args[0] if e.args else str(e)))

        message = _("%s unit(s) updated.") % updated
        if errors:
            message += "\n\n" + _("Issues:") + "\n" + "\n".join(errors)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Import Units"),
                'message': message,
                'sticky': bool(errors),
                'type': 'warning' if errors else 'success',
            },
        }
