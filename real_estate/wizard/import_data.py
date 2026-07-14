# -*- coding: utf-8 -*-
import base64
import io
import os

from odoo import models, fields, _
from odoo.exceptions import ValidationError
from datetime import datetime
root_path = os.path.dirname(os.path.abspath(__file__))

from xlrd import open_workbook
from openpyxl import load_workbook


class ImportData(models.TransientModel):
    _name = "import.data"
    _description = 'Import Data'

    file = fields.Binary(string='Xls file')

    x_action = fields.Selection([
        ('members', 'Members'),
        ('invoices', 'Invoices'),
        ('unit_inventory', 'Unit Inventory'),
        ],
        string='Action', required=True)

    def search_record(self,model,key,value):
        if value:
            record = self.env[model].search([(key, '=', value)], limit=1)
            if record:
                return record.id  
            else:
                raise ValidationError("Please create %s First in %s" %(value,model))

    def m2m_records(self,key,value,model):

        return [[6, False, [self.search_record(model, key, rec) for rec in [value]]]]

    def import_members(self):

        member_keys = [
            'name', 'agent_id', 'relation_id', 'relation_name', 'street', 'street2', 'city_id', 'state_id',
            'zip', 'country_id',
            'is_same', 'corespondence_street', 'corespondence_street2', 'corespondence_city', 'corespondence_state_id',
            'corespondence_zip',
            'corespondence_country_id', 'cnic', 'gender', 'passport', 'country_code', 'mobile', 'secondary_phone',
            'phone', 'email']
        #
        # member_keys = [
        #     'name', 'is_member', 'agent_id', 'relation_id', 'relation_name', 'street', 'street2', 'city_id', 'state_id',
        #     'zip', 'country_id',
        #     'is_same', 'corespondence_street', 'corespondence_street2', 'corespondence_city', 'corespondence_state_id',
        #     'corespondence_zip',
        #     'corespondence_country_id', 'cnic', 'gender', 'passport', 'country_code', 'mobile', 'secondary_phone',
        #     'phone', 'email', 'correspondence_mean_id', 'correnponding_option']

        # kin_keys = ['name', 'relation_id', 'relation_name','street', 'street2', 'city_id', 'state_id', 'zip', 'country_id', 
            # 'is_same', 'corespondence_street','corespondence_street2','corespondence_city','corespondence_state_id', 'corespondence_zip',
            # 'corespondence_country_id','gender','relation_with_member', 'country_code', 'mobile', 'secondary_phone', 'phone', 'cnic', 'passport', ]

        sheets = open_workbook(file_contents=base64.b64decode(self.file)).sheets()

        for sheet in sheets:
            for row in range(1,sheet.nrows):
                values = [sheet.cell(row, col).value for col in range(sheet.ncols)]
                member_values = values[:len(member_keys)]
                # kin_values = values[len(member_keys):]
                
                member_data = dict(zip(member_keys, member_values))
                # kin_data = dict(zip(kin_keys, kin_values))
                # print(member_data)
                # print(kin_data)

                member_data['agent_id'] =  self.m2m_records('name',member_data['agent_id'],'res.partner')
                member_data['city_id'] = self.search_record('city', 'name', member_data['city_id'])
                member_data['state_id'] = self.search_record('res.country.state','name', member_data['state_id'])
                member_data['country_id'] = self.search_record('res.country', 'name', member_data['country_id'])
                member_data['corespondence_city'] = self.search_record('city', 'name', member_data['corespondence_city'])
                member_data['corespondence_state_id'] = self.search_record('res.country.state','name', member_data['corespondence_state_id'])
                member_data['corespondence_country_id'] = self.search_record('res.country', 'name', member_data['corespondence_country_id'])
                member_data['country_code'] = self.search_record('res.country.code', 'name', '00'+str(member_data['country_code']).partition('.')[0])
                # member_data['correspondence_mean_id'] =  self.m2m_records('name',member_data['correspondence_mean_id'],'correnpondence.mean')

                # kin_data['city_id'] = self.search_record('city', 'name', kin_data['city_id'])
                # kin_data['state_id'] = self.search_record('res.country.state','name', kin_data['state_id'])
                # kin_data['country_id'] = self.search_record('res.country', 'name', kin_data['country_id'])
                # kin_data['corespondence_city'] = self.search_record('city', 'name', kin_data['corespondence_city'])
                # kin_data['corespondence_state_id'] = self.search_record('res.country.state','name', kin_data['corespondence_state_id'])
                # kin_data['corespondence_country_id'] = self.search_record('res.country', 'name', kin_data['corespondence_country_id'])
                # kin_data['country_code'] = self.search_record('res.country.code', 'name', '00'+str(kin_data['country_code']).partition('.')[0])

                partner = self.env['res.member']
                if not partner.search(['&',('name', '=', member_data['name']),'|',('cnic', '=', member_data['cnic']),('passport', '=', member_data['passport'])]):
                    members = self.env['res.member'].create(member_data)
                    # kin_data['partner_id'] = members.id
                    # kin = self.env['res.member'].create(kin_data)
                    self.env.cr.commit()

    def import_invoices(self):
        wb = open_workbook(file_contents=base64.b64decode(self.file))
        ws = wb.sheets()[0]
        data = {}
        dataline = {}
        for s in wb.sheets():
            for row in range(s.nrows):
                if row > 0:
                    for col in range(s.ncols):
                        if col == 0 and s.cell(row, col).value:
                            data['name'] = s.cell(row, col).value

                        elif col == 1 and s.cell(row, col).value:

                            data['date'] = s.cell(row, col).value
                            recp = self.env['import.invoices'].create(data)

                        elif col == 2 and s.cell(row, col).value:

                            dataline['member_id'] = self.env['res.member'].search([('ref', '=', s.cell(row, col).value.rstrip())]).id
                        elif col == 3 and s.cell(row, col).value:

                            dataline['contract_id'] = self.env['res.contract'].search([('name', '=', s.cell(row, col).value.rstrip())]).id
                        elif col == 4 and s.cell(row, col).value:

                            dataline['ammenities_id'] = self.env['amenity.amenity'].search([('name', '=', s.cell(row, col).value.rstrip())]).id
                        elif col == 5 and s.cell(row, col).value:

                            dataline['meter_number'] = s.cell(row, col).value
                        elif col == 6 and s.cell(row, col).value:

                            dataline['previous_reading'] = s.cell(row, col).value
                        elif col == 7 and s.cell(row, col).value:

                            dataline['new_reading'] = s.cell(row, col).value
                        elif col == 8 and s.cell(row, col).value:

                            dataline['unit_rate'] = s.cell(row, col).value
                        elif col == 9 and s.cell(row, col).value:

                            dataline['amount'] = s.cell(row, col).value
                            dataline['import_invices_id'] = recp.id

                            recp.import_invoice_line.create(dataline)

    # ------------------------------------------------------------------
    # Unit Inventory (plot.inventory) import from .xlsx
    # ------------------------------------------------------------------

    UNIT_INVENTORY_HEADERS = {
        'serial number': 'serial_number',
        'plot number': 'name',
        'project type': 'project_type',
        'society': 'society_id',
        'phase': 'phase_id',
        'sector': 'sector_id',
        'street': 'street_id',
        'category': 'category_id',
        'unit category type': 'unit_category_type_id',
        'product': 'unit_category_type_id',
        'size': 'size_id',
        'unit class': 'unit_class_id',
        'type': 'unit_class_id',
        'net area': 'net_area',
        'balcony area': 'balcony_area',
        'standard area (sft)': 'standard_area',
        'total area (sft)': 'standard_area',
        'actual area': 'actual_area',
        'state': 'state',
        'possession status': 'possession_status',
    }

    def _find_by_name(self, model, name, domain=None, row_no=0):
        if not name:
            return False
        name = str(name).strip()
        record = self.env[model].search([('name', '=', name)] + (domain or []), limit=1)
        if not record:
            raise ValidationError(
                _("Row %s: '%s' not found in %s. Please create it first.") % (row_no, name, model))
        return record.id

    def _selection_value(self, field_name, label, row_no=0):
        if not label:
            return False
        label = str(label).strip()
        for value, string in self.env['plot.inventory']._fields[field_name].selection:
            if label.lower() in (value.lower(), string.lower()):
                return value
        raise ValidationError(
            _("Row %s: '%s' is not a valid value for %s.") % (row_no, label, field_name))

    def import_unit_inventory(self):
        wb = load_workbook(io.BytesIO(base64.b64decode(self.file)), read_only=True, data_only=True)
        Inventory = self.env['plot.inventory']
        created = updated = 0

        for sheet in wb.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            columns = [self.UNIT_INVENTORY_HEADERS.get(str(h or '').strip().lower()) for h in header]

            for row_no, row in enumerate(rows, start=2):
                data = {field: row[col] for col, field in enumerate(columns) if field and col < len(row)}
                if not any(v not in (None, '') for v in data.values()):
                    continue

                society_id = self._find_by_name(
                    'society', data.get('society_id'), [('is_society', '=', True)], row_no)
                phase_id = self._find_by_name(
                    'society', data.get('phase_id'),
                    [('is_society', '!=', True), ('society_id', '=', society_id)], row_no)
                sector_id = self._find_by_name(
                    'sector', data.get('sector_id'), [('phase_id', '=', phase_id)], row_no)
                street_id = self._find_by_name(
                    'street', data.get('street_id'), [('sector_id', '=', sector_id)], row_no)

                vals = {
                    'name': str(data['name']).strip() if data.get('name') else False,
                    'project_type': self._selection_value(
                        'project_type', data.get('project_type'), row_no) or 'housing_society',
                    'society_id': society_id,
                    'phase_id': phase_id,
                    'sector_id': sector_id,
                    'street_id': street_id,
                    'category_id': self._find_by_name('plot.category', data.get('category_id'), row_no=row_no),
                    'unit_category_type_id': self._find_by_name(
                        'unit.category.type', data.get('unit_category_type_id'), row_no=row_no),
                    'unit_class_id': self._find_by_name('unit.class', data.get('unit_class_id'), row_no=row_no),
                    'size_id': self._find_by_name('unit.size', data.get('size_id'), row_no=row_no),
                    'net_area': float(data.get('net_area') or 0),
                    'balcony_area': float(data.get('balcony_area') or 0),
                    'standard_area': float(data.get('standard_area') or 0),
                    'actual_area': float(data.get('actual_area') or 0),
                    'state': self._selection_value('state', data.get('state'), row_no) or 'avalible_for_sale',
                    'possession_status': self._selection_value(
                        'possession_status', data.get('possession_status'), row_no) or 'pending',
                }

                serial = str(data['serial_number']).strip() if data.get('serial_number') else False
                existing = serial and Inventory.search([('serial_number', '=', serial)], limit=1)
                if existing:
                    existing.write(vals)
                    updated += 1
                else:
                    if serial:
                        vals['serial_number'] = serial
                    elif not (sector_id and vals['category_id'] and vals['unit_category_type_id']):
                        raise ValidationError(
                            _("Row %s: Sector, Category and Product are required to generate the serial number.") % row_no)
                    Inventory.create(vals)
                    created += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Unit Inventory Import'),
                'message': _('%s record(s) created, %s record(s) updated.') % (created, updated),
                'type': 'success',
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }